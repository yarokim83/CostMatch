import customtkinter as ctk
from tkinter import filedialog, messagebox
import pandas as pd
import os
import threading
import openpyxl
from openpyxl.styles import PatternFill

# Set theme
ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("blue")

class ExcelComparatorApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        # Window setup
        self.title("CostMatch")
        self.geometry("1000x800")
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(2, weight=1) # Result area expands

        # Variables
        self.file1_list = [] # List of dictionaries: {'type': str, 'path': str, 'widgets': dict}
        self.file2_path = None
        self.dept_types = ['QC', 'ARMGC', 'RS', 'YT', 'YC', 'ECH', 'FL', '기타', 'Lubricants']

        # UI Elements
        self.create_widgets()

    def create_widgets(self):
        # Header
        self.header_frame = ctk.CTkFrame(self, corner_radius=0)
        self.header_frame.grid(row=0, column=0, sticky="ew", padx=0, pady=(0, 10))
        self.header_label = ctk.CTkLabel(self.header_frame, text="CostMatch", font=ctk.CTkFont(size=24, weight="bold"))
        self.header_label.pack(pady=10)

        # Main Content Area
        self.main_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.main_frame.grid(row=1, column=0, sticky="nsew", padx=20, pady=5)
        self.main_frame.grid_columnconfigure(0, weight=1)

        # --- Section 1: Cost Settlement File (File 2) ---
        self.file2_frame = ctk.CTkFrame(self.main_frame)
        self.file2_frame.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        self.file2_frame.grid_columnconfigure(1, weight=1)
        
        self.file2_btn = ctk.CTkButton(self.file2_frame, text="파일 2 (비용 정산)", command=self.select_file2, width=150, fg_color="#E53935", hover_color="#D32F2F")
        self.file2_btn.grid(row=0, column=0, padx=10, pady=10)
        self.file2_label = ctk.CTkLabel(self.file2_frame, text="선택된 파일 없음", text_color="gray", anchor="w")
        self.file2_label.grid(row=0, column=1, sticky="ew", padx=10)

        # --- Section 2: Department Files (File 1 List) ---
        self.file1_frame = ctk.CTkFrame(self.main_frame)
        self.file1_frame.grid(row=1, column=0, sticky="nsew", pady=(0, 10))
        self.file1_frame.grid_columnconfigure(0, weight=1)
        
        # Title & Add Button
        self.f1_header = ctk.CTkFrame(self.file1_frame, fg_color="transparent")
        self.f1_header.pack(fill="x", padx=10, pady=5)
        
        ctk.CTkLabel(self.f1_header, text="부서별 파일 (File 1)", font=ctk.CTkFont(size=16, weight="bold")).pack(side="left")
        ctk.CTkButton(self.f1_header, text="+ 파일 추가", command=self.add_file_row, width=100).pack(side="right", padx=5)
        ctk.CTkButton(self.f1_header, text="++ 다중 추가", command=self.add_multiple_files, width=100, fg_color="#43A047", hover_color="#2E7D32").pack(side="right", padx=5)
        ctk.CTkButton(self.f1_header, text="📂 폴더 선택", command=self.add_files_from_folder, width=100, fg_color="#1E88E5", hover_color="#1565C0").pack(side="right", padx=5)

        # Scrollable List
        self.file_list_frame = ctk.CTkScrollableFrame(self.file1_frame, height=200)
        self.file_list_frame.pack(fill="both", expand=True, padx=10, pady=5)
        
        # Add initial row (default QC)
        self.add_file_row(default_type='QC')

        # --- Section 3: Result Area ---
        self.result_text = ctk.CTkTextbox(self, width=800, height=300, font=ctk.CTkFont(family="Consolas", size=12))
        self.result_text.grid(row=2, column=0, sticky="nsew", padx=20, pady=10)
        self.result_text.insert("0.0", "파일 2(비용 정산)와 각 부서별 파일(File 1)을 선택하고 비교 버튼을 눌러주세요.\n")
        self.result_text.configure(state="disabled")

        # --- Section 4: Action Buttons ---
        self.action_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.action_frame.grid(row=3, column=0, sticky="ew", padx=20, pady=20)
        
        self.compare_btn = ctk.CTkButton(self.action_frame, text="비교 분석 시작", command=self.start_comparison, height=50, font=ctk.CTkFont(size=18, weight="bold"))
        self.compare_btn.pack(fill="x")

    def add_file_row(self, default_type='QC', file_path=None):
        row_idx = len(self.file1_list)
        
        row_frame = ctk.CTkFrame(self.file_list_frame, fg_color="transparent")
        row_frame.pack(fill="x", pady=2)
        
        # Type Dropdown
        type_var = ctk.StringVar(value=default_type)
        type_menu = ctk.CTkOptionMenu(row_frame, values=self.dept_types, variable=type_var, width=120)
        type_menu.pack(side="left", padx=5)
        
        # File Label
        label_text = os.path.basename(file_path) if file_path else "선택된 파일 없음"
        label_color = "white" if file_path else "gray"
        path_label = ctk.CTkLabel(row_frame, text=label_text, text_color=label_color, width=300, anchor="w")
        path_label.pack(side="left", padx=5, fill="x", expand=True)
        
        # Select Button
        # Use closure to capture current row context
        def select_cmd(lbl=path_label, idx=row_idx):
            self.select_file1_for_row(lbl, idx)
            
        sel_btn = ctk.CTkButton(row_frame, text="선택", command=select_cmd, width=60)
        sel_btn.pack(side="left", padx=5)
        
        # Remove Button
        def remove_cmd(frm=row_frame, idx=row_idx):
            self.remove_file_row(frm, idx)
            
        del_btn = ctk.CTkButton(row_frame, text="X", command=remove_cmd, width=30, fg_color="#D32F2F", hover_color="#B71C1C")
        del_btn.pack(side="left", padx=5)
        
        # Store widget references
        self.file1_list.append({
            'frame': row_frame,
            'type_var': type_var,
            'path_label': path_label,
            'path': file_path,
            'id': row_idx # Unique ID for this row
        })

    def add_multiple_files(self):
        filenames = filedialog.askopenfilenames(filetypes=[("Excel files", "*.xlsx *.xls *.xlsm")])
        if filenames:
            for filename in filenames:
                # Guess type from filename
                guessed_type = 'QC' # Default
                fname_upper = os.path.basename(filename).upper()
                
                # Check for MV or TL first
                if 'LUB' in fname_upper or 'LUBRICANT' in fname_upper:
                    guessed_type = 'Lubricants'
                elif 'MV' in fname_upper or 'TL' in fname_upper:
                    guessed_type = '기타'
                # Check for TC (maps to ARMGC)
                elif 'TC' in fname_upper:
                    guessed_type = 'ARMGC'
                else:
                    for dtype in self.dept_types:
                        if dtype in fname_upper:
                            guessed_type = dtype
                            break
                
                self.add_file_row(default_type=guessed_type, file_path=filename)
            
            self.log(f"[설정] {len(filenames)}개 파일이 추가되었습니다.")

    def add_files_from_folder(self):
        folder_path = filedialog.askdirectory()
        if folder_path:
            self.log(f"\n[폴더 스캔] 경로: {folder_path}")
            
            # Remove empty rows first
            empty_rows = [item for item in self.file1_list if item['path'] is None]
            for item in empty_rows:
                self.remove_file_row(item['frame'], item['id'])
            
            count = 0
            for filename in os.listdir(folder_path):
                if filename.lower().endswith(('.xlsx', '.xls', '.xlsm')):
                    full_path = os.path.join(folder_path, filename)
                    
                    # Guess type
                    guessed_type = None
                    fname_upper = filename.upper()
                    
                    # Debug log for each file
                    # self.log(f"  - 파일 검사: {filename}")
                    
                    # Check for MV or TL first
                    if 'LUB' in fname_upper or 'LUBRICANT' in fname_upper:
                        guessed_type = 'Lubricants'
                    elif 'MV' in fname_upper or 'TL' in fname_upper:
                        guessed_type = '기타'
                    # Check for TC (maps to ARMGC)
                    elif 'TC' in fname_upper:
                        guessed_type = 'ARMGC'
                    else:
                        for dtype in self.dept_types:
                            if dtype in fname_upper:
                                guessed_type = dtype
                                break
                    
                    if guessed_type:
                        self.add_file_row(default_type=guessed_type, file_path=full_path)
                        count += 1
                        self.log(f"    -> [추가됨] 타입: {guessed_type} | 파일: {filename}")
                    else:
                        self.log(f"    -> [무시됨] 부서명 매칭 실패: {filename}")
            
            if count > 0:
                self.log(f"[설정] 폴더에서 {count}개 파일을 자동으로 추가했습니다.")
            else:
                messagebox.showinfo("알림", "해당 폴더에서 부서명(QC, ARMGC 등)이 포함된 엑셀 파일을 찾을 수 없습니다.")
                # If no files added, ensure at least one empty row exists
                if not self.file1_list:
                    self.add_file_row()

    def remove_file_row(self, frame, row_id):
        frame.destroy()
        # Remove from list based on ID
        self.file1_list = [item for item in self.file1_list if item['id'] != row_id]

    def select_file1_for_row(self, label_widget, row_id):
        filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls *.xlsm")])
        if filename:
            # Find the item in list
            for item in self.file1_list:
                if item['id'] == row_id:
                    item['path'] = filename
                    label_widget.configure(text=os.path.basename(filename), text_color="white")
                    self.log(f"[설정] {item['type_var'].get()} 파일 선택됨: {filename}")
                    break

    def select_file1(self):
        pass # Deprecated

    def select_file2(self):
        filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls *.xlsm")])
        if filename:
            self.file2_path = filename
            self.file2_label.configure(text=os.path.basename(filename), text_color="white")
            self.log(f"[설정] 파일 2 선택됨: {filename}")

    def log(self, message):
        self.result_text.configure(state="normal")
        self.result_text.insert("end", message + "\n")
        self.result_text.see("end")
        self.result_text.configure(state="disabled")
        
        # Save to file
        with open("analysis_log.txt", "a", encoding="utf-8") as f:
            f.write(message + "\n")

    def clear_log(self):
        self.result_text.configure(state="normal")
        self.result_text.delete("1.0", "end")
        self.result_text.configure(state="disabled")
        
        # Clear log file
        with open("analysis_log.txt", "w", encoding="utf-8") as f:
            f.write("")

    def start_comparison(self):
        # Check File 2
        if not self.file2_path:
            messagebox.showwarning("경고", "파일 2(비용 정산)를 선택해주세요.")
            return

        # Check File 1 List (at least one valid file)
        valid_files = [f for f in self.file1_list if f['path']]
        if not valid_files:
            messagebox.showwarning("경고", "부서별 파일(File 1)을 적어도 하나 이상 선택해주세요.")
            return

        self.compare_btn.configure(state="disabled", text="분석 중...")
        self.clear_log()
        thread = threading.Thread(target=self.run_analysis)
        thread.start()

    def load_excel_smart(self, filepath, required_columns):
        """
        Try to find the header row that contains the required columns.
        """
        # Read first few rows to inspect
        try:
            # First, try reading normally
            df = pd.read_excel(filepath)
            if all(col in df.columns for col in required_columns):
                return df
            
            # If not found, try to find the header in the first 20 rows
            df_raw = pd.read_excel(filepath, header=None, nrows=20)
            
            header_row_idx = -1
            for idx, row in df_raw.iterrows():
                # Check if this row contains ALL required columns (fuzzy match or exact)
                # We convert row values to string and check
                row_values = [str(v).strip() for v in row.values if pd.notna(v)]
                
                # Check if all required columns are present in this row
                if all(req in row_values for req in required_columns):
                    header_row_idx = idx
                    break
            
            if header_row_idx != -1:
                self.log(f"  -> {os.path.basename(filepath)}: 헤더를 {header_row_idx+1}행에서 찾았습니다.")
                return pd.read_excel(filepath, header=header_row_idx)
            
            # If still not found, return original to let the caller handle the error
            return df
            
        except Exception as e:
            raise e

    def run_analysis(self):
        try:
            self.log(">>> 데이터 로딩 및 분석 시작...")
            
            # Check File 2
            if not self.file2_path:
                self.log("\n[오류] 파일 2(비용 정산)가 선택되지 않았습니다.")
                return

            # Check File 1 List
            valid_files = [f for f in self.file1_list if f['path']]
            if not valid_files:
                self.log("\n[오류] 선택된 부서별 파일(File 1)이 없습니다.")
                return

            # Deduplicate by path
            unique_files = {}
            for f in valid_files:
                if f['path'] not in unique_files:
                    unique_files[f['path']] = f
            valid_files = list(unique_files.values())

            # Sort files by predefined order
            # Order: QC -> ARMGC -> RS -> YT -> YC -> ECH -> FL -> 기타 -> Lubricants
            order_map = {val: i for i, val in enumerate(self.dept_types)}
            valid_files.sort(key=lambda x: order_map.get(x['type_var'].get(), 99))

            # --- Load File 2 (Cost) ---
            # We will try to find '발주처' or similar
            # Added 'PR', 'Account Cod', '담당자' to request, but we'll handle them if missing
            req_cols_2 = ['Account name', '발주금액', 'PR No.'] 
            # We don't enforce others in req_cols_2 initially to avoid errors if they are named differently.
            
            self.log(f"파일 2 로드 중: {os.path.basename(self.file2_path)}")
            df2 = self.load_excel_smart(self.file2_path, req_cols_2)
            
            # Helper to find column with candidates
            def find_col(df, candidates, default_name):
                for cand in candidates:
                    if cand in df.columns:
                        return cand
                # If not found, create empty
                df[default_name] = ""
                return default_name

            def filter_account_name(df, account_names):
                if isinstance(account_names, str):
                    account_names = [account_names]
                normalized_account = df['Account name'].astype(str).str.replace(" ", "", regex=False).str.strip()
                normalized_targets = [str(name).replace(" ", "").strip() for name in account_names]
                return df[normalized_account.isin(normalized_targets)].copy()

            # Find '발주처' column
            vendor_col_f2 = find_col(df2, ['발주처', 'Vendor', 'Supplier', '거래처', 'Vendor Name'], '발주처')
            
            # Find 'PR' column (separate from PR No.)
            pr_short_col_f2 = find_col(df2, ['PR'], 'PR')
            
            # Find 'Account Cod' column
            acc_code_col_f2 = find_col(df2, ['Account Cod', 'Account Code', 'Account No', 'Account No.'], 'Account Cod')
            
            # Find '담당자' column
            manager_col_f2 = find_col(df2, ['담당자', 'Manager', 'Person in charge'], '담당자')

            self.log(f"  -> 파일 2 추가 컬럼 확인: 발주처='{vendor_col_f2}', PR='{pr_short_col_f2}', Account Cod='{acc_code_col_f2}', 담당자='{manager_col_f2}'")
            
            # Account Mapping (Material)
            account_mapping_mat = {
                'QC': '장비 자재비-QC',
                'ARMGC': '장비 자재비-ATC', # Mapped to ATC
                'RS': '장비 자재비-RS',
                'YT': '장비 자재비-YT',
                'YC': '장비 자재비-YC',
                'ECH': '장비 자재비-ECH',
                'FL': '장비 자재비-FL',
                '기타': '장비 자재비-기타',
                'Lubricants': '동력비-윤활유'
            }

            # Account Mapping (Outsourcing)
            account_mapping_out = {
                'QC': ['수선유지비-외주수리-QC', '수선유지비-외주수리 QC'],
                'ARMGC': ['수선유지비-외주수리-ATC', '수선유지비-외주수리-ARMGC', '수선유지비-외주수리-TC'], # Mapped to ATC
                'RS': '수선유지비-외주수리-RS',
                'YT': '수선유지비-외주수리-YT',
                'YC': '수선유지비-외주수리-YC',
                'ECH': '수선유지비-외주수리-ECH',
                'FL': '수선유지비-외주수리-FL',
                '기타': '수선유지비-외주수리-기타',
                'Lubricants': '수선유지비-외주수리-Lubricants'
            }

            # --- Process Each File 1 ---
            final_material_rows_left = []
            final_material_rows_right = []
            final_outsourcing_rows_left = []
            final_outsourcing_rows_right = []
            
            missing_doc_nos = set() # For highlighting later (Material)
            missing_doc_nos_out = set() # For highlighting later (Outsourcing)
            
            # Fix: Track MATCHED PR Nos instead of unmatched.
            # If we track unmatched, a PR might be flagged as unmatched in one file loop 
            # but matched in another (if multiple files per dept), causing false positives.
            matched_pr_nos_mat = set()
            matched_pr_nos_out = set()
            
            # Track amount mismatches (PR No. where Total Price != 발주금액)
            amount_mismatch_pr_nos_mat = set()
            amount_mismatch_pr_nos_out = set()
            
            report_cols = [
                "Type", "Date", "Part No.", "Part Type", "Part Group", 
                "Description", "Qty", "Unit Price", "Total Price", 
                "Doc No.", "Mach No.", "Vendor"
            ]
            # Updated report columns for File 2
            report_cols_f2 = ['Account name', 'Account Cod', 'PR No.', 'PR', '발주처', '발주금액', '담당자']

            for file_item in valid_files:
                f_type = file_item['type_var'].get()
                f_path = file_item['path']
                
                self.log(f"\n>>> [{f_type}] 파일 처리 중: {os.path.basename(f_path)}")
                
                # 1. Filter File 2 for this type (Material)
                target_account_mat = account_mapping_mat.get(f_type, f'장비 자재비-{f_type}')
                df2_mat = filter_account_name(df2, target_account_mat)
                
                # 2. Filter File 2 for this type (Outsourcing)
                target_account_out = account_mapping_out.get(f_type, f'수선유지비-외주수리-{f_type}')
                df2_out = filter_account_name(df2, target_account_out)
                
                if df2_mat.empty:
                    self.log(f"  [주의] 파일 2에서 자재 계정 '{target_account_mat}' 항목을 찾을 수 없습니다.")
                else:
                    self.log(f"  -> 파일 2 필터링 (자재): '{target_account_mat}' ({len(df2_mat)} 건)")

                if df2_out.empty:
                    self.log(f"  [주의] 파일 2에서 외주수리 계정 '{target_account_out}' 항목을 찾을 수 없습니다.")
                else:
                    self.log(f"  -> 파일 2 필터링 (외주수리): '{target_account_out}' ({len(df2_out)} 건)")


                # 2. Prepare File 2 Data (Group by PR No.)
                # Find PR No. column
                pr_col = None
                if 'PR No..1' in df2.columns:
                    pr_col = 'PR No..1'
                elif 'PR No.' in df2.columns:
                    pr_col = 'PR No.'
                else:
                    candidates = [c for c in df2.columns if 'PR' in str(c) and 'No' in str(c)]
                    if candidates:
                        pr_col = candidates[0]
                
                # Create grouped DataFrames for Material
                df2_mat_grouped = pd.DataFrame()
                if pr_col and pr_col in df2_mat.columns:
                    df2_mat[pr_col] = df2_mat[pr_col].astype(str).str.strip()
                    # Ensure 'PR No.' column exists for matching
                    if pr_col != 'PR No.':
                        df2_mat['PR No.'] = df2_mat[pr_col]
                    else:
                        # If pr_col is 'PR No.', ensure it's clean
                        df2_mat['PR No.'] = df2_mat[pr_col]
                # Create grouped DataFrames for Material
                df2_mat_grouped = pd.DataFrame()
                if pr_col and pr_col in df2_mat.columns:
                    df2_mat[pr_col] = df2_mat[pr_col].astype(str).str.strip()
                    # Group by PR No. and sum Amount, but we also need to keep other columns like Account name and Vendor.
                    # Since we are grouping, non-numeric columns might be lost or need aggregation.
                    # For the report, if we have multiple rows with same PR No, what should we show for Vendor?
                    # Usually they are the same. We can take 'first'.
                    
                    agg_dict = {'발주금액': 'sum'}
                    # Add other columns to aggregation (taking first value)
                    for col in [vendor_col_f2, pr_short_col_f2, acc_code_col_f2, manager_col_f2]:
                        if col:
                            agg_dict[col] = 'first'
                    agg_dict['Account name'] = 'first'
                        
                    df2_mat_grouped = df2_mat.groupby(pr_col).agg(agg_dict).reset_index()
                    
                    if pr_col != 'PR No.':
                        df2_mat_grouped = df2_mat_grouped.rename(columns={pr_col: 'PR No.'})
                    
                    # Rename cols to standard names if needed
                    rename_map = {}
                    if vendor_col_f2 and vendor_col_f2 != '발주처': rename_map[vendor_col_f2] = '발주처'
                    if pr_short_col_f2 and pr_short_col_f2 != 'PR': rename_map[pr_short_col_f2] = 'PR'
                    if acc_code_col_f2 and acc_code_col_f2 != 'Account Cod': rename_map[acc_code_col_f2] = 'Account Cod'
                    if manager_col_f2 and manager_col_f2 != '담당자': rename_map[manager_col_f2] = '담당자'
                    
                    if rename_map:
                        df2_mat_grouped = df2_mat_grouped.rename(columns=rename_map)
                        
                else:
                    df2_mat_grouped = pd.DataFrame(columns=['PR No.', '발주금액', '발주처', 'PR', 'Account Cod', '담당자', 'Account name'])
                    if not df2_mat.empty:
                        self.log(f"  -> ⚠️ 'PR No.' 관련 컬럼을 찾을 수 없어 자재 매칭이 불가능할 수 있습니다.")
                
                # Create grouped DataFrames for Outsourcing
                df2_out_grouped = pd.DataFrame()
                if pr_col and pr_col in df2_out.columns:
                    df2_out[pr_col] = df2_out[pr_col].astype(str).str.strip()
                # Create grouped DataFrames for Outsourcing
                df2_out_grouped = pd.DataFrame()
                if pr_col and pr_col in df2_out.columns:
                    df2_out[pr_col] = df2_out[pr_col].astype(str).str.strip()
                    # Ensure 'PR No.' column exists for matching
                    if pr_col != 'PR No.':
                        df2_out['PR No.'] = df2_out[pr_col]
                    else:
                        df2_out['PR No.'] = df2_out[pr_col]
                    
                    agg_dict = {'발주금액': 'sum'}
                    # Add other columns to aggregation
                    for col in [vendor_col_f2, pr_short_col_f2, acc_code_col_f2, manager_col_f2]:
                        if col:
                            agg_dict[col] = 'first'
                    agg_dict['Account name'] = 'first'

                    df2_out_grouped = df2_out.groupby(pr_col).agg(agg_dict).reset_index()
                    
                    if pr_col != 'PR No.':
                        df2_out_grouped = df2_out_grouped.rename(columns={pr_col: 'PR No.'})
                        
                    # Rename cols to standard names
                    rename_map = {}
                    if vendor_col_f2 and vendor_col_f2 != '발주처': rename_map[vendor_col_f2] = '발주처'
                    if pr_short_col_f2 and pr_short_col_f2 != 'PR': rename_map[pr_short_col_f2] = 'PR'
                    if acc_code_col_f2 and acc_code_col_f2 != 'Account Cod': rename_map[acc_code_col_f2] = 'Account Cod'
                    if manager_col_f2 and manager_col_f2 != '담당자': rename_map[manager_col_f2] = '담당자'
                    
                    if rename_map:
                        df2_out_grouped = df2_out_grouped.rename(columns=rename_map)
                else:
                    df2_out_grouped = pd.DataFrame(columns=['PR No.', '발주금액', '발주처', 'PR', 'Account Cod', '담당자', 'Account name'])
                    if not df2_out.empty:
                        self.log(f"  -> ⚠️ 'PR No.' 관련 컬럼을 찾을 수 없어 외주수리 매칭이 불가능할 수 있습니다.")
                
                # 3. Load File 1
                req_cols_1 = ['Doc No.', 'Part Group', 'Total Price', 'Part No.', 'Vendor']
                df1 = self.load_excel_smart(f_path, req_cols_1)
                df1.columns = [str(col).strip() for col in df1.columns]
                
                # --- Filter Data ---
                # Material (WIRE ROPE, INVENTORY, TIRE)
                target_groups_mat = ['WIRE ROPE', 'INVENTORY', 'TIRE']
                part_group_upper = df1['Part Group'].astype(str).str.upper()
                mask_lub = part_group_upper.str.contains('LUB', na=False)
                if f_type == 'Lubricants':
                    mask_mat = mask_lub
                else:
                    mask_mat = part_group_upper.isin(target_groups_mat) & ~mask_lub
                df1_mat = df1[mask_mat].copy()
                if f_type == 'Lubricants' and not df1_mat.empty and 'Type' in df1_mat.columns:
                    df1_mat['Type'] = 'Lubricants'
                
                # Outsourcing (Contains 'OUTSOURCING')
                if 'Part Group' in df1.columns:
                    mask_out = df1['Part Group'].astype(str).str.upper().str.contains('OUTSOURCING', na=False)
                    df1_out = df1[mask_out].copy()
                    self.log(f"  -> File 1 외주수리 필터링: {len(df1_out)} 건")
                else:
                    df1_out = pd.DataFrame()
                    self.log("  -> ⚠️ File 1에서 'Part Group' 컬럼을 찾을 수 없어 외주수리 추출이 불가능합니다.")

                # --- Comparison (for Highlighting) ---
                # 1. Material (Doc No. vs PR No.)
                df1_mat['Doc No.'] = df1_mat['Doc No.'].astype(str).str.strip()
                df1_mat_grouped = df1_mat.groupby('Doc No.')['Total Price'].sum().reset_index()
                
                merged = pd.merge(
                    df1_mat_grouped, 
                    df2_mat_grouped, 
                    left_on='Doc No.', 
                    right_on='PR No.', 
                    how='left', 
                    indicator=True
                )
                
                only_in_f1 = merged[merged['_merge'] == 'left_only']
                missing_doc_nos.update(only_in_f1['Doc No.'].tolist())
                
                only_in_f1 = merged[merged['_merge'] == 'left_only']
                missing_doc_nos.update(only_in_f1['Doc No.'].tolist())
                
                # Identify MATCHED PR Nos and check for amount mismatches
                if not df2_mat.empty:
                    doc_nos_in_group = set(df1_mat['Doc No.'].tolist())
                    # Create a dict of Doc No -> Total Price sum for this group
                    doc_no_amounts = df1_mat.groupby('Doc No.')['Total Price'].sum().to_dict()
                    
                    # Check which PRs in this File 2 slice match ANY Doc No in this File 1
                    for idx, row in df2_mat.iterrows():
                        # Use 'PR No.' which we standardized above
                        pr_val = str(row.get('PR No.', '')).strip()
                        if pr_val and pr_val in doc_nos_in_group:
                            matched_pr_nos_mat.add(pr_val)
                            
                            # Check if amounts match
                            f1_amount = doc_no_amounts.get(pr_val, 0)
                            f2_amount = row.get('발주금액', 0)
                            # Convert to float for comparison
                            try:
                                f1_amount = float(f1_amount) if f1_amount else 0
                                f2_amount = float(f2_amount) if f2_amount else 0
                                # Compare with small tolerance for floating point
                                if abs(f1_amount - f2_amount) > 0.01:
                                    amount_mismatch_pr_nos_mat.add(pr_val)
                            except (ValueError, TypeError):
                                pass  # Skip if conversion fails

                # 2. Outsourcing (Doc No. vs PR No.)
                if not df1_out.empty:
                    df1_out['Doc No.'] = df1_out['Doc No.'].astype(str).str.strip()
                    df1_out_grouped = df1_out.groupby('Doc No.')['Total Price'].sum().reset_index()
                    
                    # Use df2_out_grouped (grouped by PR No.) for Outsourcing
                    merged_out = pd.merge(
                        df1_out_grouped,
                        df2_out_grouped,
                        left_on='Doc No.',
                        right_on='PR No.', 
                        how='left', 
                        indicator=True
                    )
                    
                    only_in_f1_out = merged_out[merged_out['_merge'] == 'left_only']
                    missing_doc_nos_out.update(only_in_f1_out['Doc No.'].tolist())

                    only_in_f1_out = merged_out[merged_out['_merge'] == 'left_only']
                    missing_doc_nos_out.update(only_in_f1_out['Doc No.'].tolist())

                    # Identify MATCHED PR Nos (Outsourcing) and check for amount mismatches
                    if not df2_out.empty:
                        doc_nos_in_group_out = set(df1_out['Doc No.'].tolist())
                        # Create a dict of Doc No -> Total Price sum for this group
                        doc_no_amounts_out = df1_out.groupby('Doc No.')['Total Price'].sum().to_dict()
                        
                        for idx, row in df2_out.iterrows():
                            # Use 'PR No.' which we standardized above
                            pr_val = str(row.get('PR No.', '')).strip()
                            if pr_val and pr_val in doc_nos_in_group_out:
                                matched_pr_nos_out.add(pr_val)
                                
                                # Check if amounts match
                                f1_amount = doc_no_amounts_out.get(pr_val, 0)
                                f2_amount = row.get('발주금액', 0)
                                # Convert to float for comparison
                                try:
                                    f1_amount = float(f1_amount) if f1_amount else 0
                                    f2_amount = float(f2_amount) if f2_amount else 0
                                    # Compare with small tolerance for floating point
                                    if abs(f1_amount - f2_amount) > 0.01:
                                        amount_mismatch_pr_nos_out.add(pr_val)
                                except (ValueError, TypeError):
                                    pass  # Skip if conversion fails

                # --- Prepare Report Data (Side-by-Side) ---
                # Material
                rows_left = []
                if not df1_mat.empty:
                    existing_cols = [c for c in report_cols if c in df1_mat.columns]
                    rows_left = df1_mat[existing_cols].to_dict('records')
                
                rows_right = []
                if not df2_mat.empty:
                    # Use df2_mat_grouped if available, but wait, df2_mat was filtered by account.
                    # The requirement is to show File 2 data. 
                    # If we use df2_mat (raw), it has all rows.
                    # If we use df2_mat_grouped, it has unique PRs.
                    # Usually for comparison we use grouped, but for listing we might want raw?
                    # The user said "extract PR / Vendor / Account Name".
                    # Let's use the raw df2_mat but ensure we have the columns.
                    
                    # Ensure all columns exist in df2_mat with standard names
                    col_map = {
                        vendor_col_f2: '발주처',
                        pr_short_col_f2: 'PR',
                        acc_code_col_f2: 'Account Cod',
                        manager_col_f2: '담당자'
                    }
                    for src, dst in col_map.items():
                        if src and src != dst and src in df2_mat.columns:
                            df2_mat[dst] = df2_mat[src]
                        elif dst not in df2_mat.columns:
                            df2_mat[dst] = ""

                    # Ensure PR No. is standard
                    if pr_col and pr_col != 'PR No.' and pr_col in df2_mat.columns:
                        df2_mat['PR No.'] = df2_mat[pr_col]
                    elif 'PR No.' not in df2_mat.columns:
                        df2_mat['PR No.'] = ""

                    cols_to_use = [c for c in report_cols_f2 if c in df2_mat.columns]
                    rows_right = df2_mat[cols_to_use].to_dict('records')

                # Align Rows
                max_len = max(len(rows_left), len(rows_right))
                while len(rows_left) < max_len: rows_left.append({c: "" for c in report_cols})
                while len(rows_right) < max_len: rows_right.append({c: "" for c in report_cols_f2})
                
                # Summary Row
                qty_sum = df1_mat['Qty'].sum() if not df1_mat.empty and 'Qty' in df1_mat.columns else 0
                price_sum = df1_mat['Total Price'].sum() if not df1_mat.empty and 'Total Price' in df1_mat.columns else 0
                sum_row_left = {c: "" for c in report_cols}
                sum_row_left['Description'] = f"{f_type}(자재)"
                sum_row_left['Qty'] = qty_sum
                sum_row_left['Total Price'] = price_sum
                rows_left.append(sum_row_left)
                
                price_sum_f2 = df2_mat['발주금액'].sum() if not df2_mat.empty and '발주금액' in df2_mat.columns else 0
                sum_row_right = {c: "" for c in report_cols_f2}
                sum_row_right['Account name'] = "Total"
                sum_row_right['발주금액'] = price_sum_f2
                rows_right.append(sum_row_right)

                # Add 2 Empty Rows
                for _ in range(2):
                    rows_left.append({c: "" for c in report_cols})
                    rows_right.append({c: "" for c in report_cols_f2})

                final_material_rows_left.extend(rows_left)
                final_material_rows_right.extend(rows_right)

                # Outsourcing
                rows_left_out = []
                if not df1_out.empty:
                    existing_cols = [c for c in report_cols if c in df1_out.columns]
                    rows_left_out = df1_out[existing_cols].to_dict('records')
                
                rows_right_out = []
                if not df2_out.empty:
                    # Ensure all columns exist
                    col_map = {
                        vendor_col_f2: '발주처',
                        pr_short_col_f2: 'PR',
                        acc_code_col_f2: 'Account Cod',
                        manager_col_f2: '담당자'
                    }
                    for src, dst in col_map.items():
                        if src and src != dst and src in df2_out.columns:
                            df2_out[dst] = df2_out[src]
                        elif dst not in df2_out.columns:
                            df2_out[dst] = ""

                    # Ensure PR No.
                    if pr_col and pr_col != 'PR No.' and pr_col in df2_out.columns:
                        df2_out['PR No.'] = df2_out[pr_col]
                    elif 'PR No.' not in df2_out.columns:
                        df2_out['PR No.'] = ""

                    cols_to_use = [c for c in report_cols_f2 if c in df2_out.columns]
                    rows_right_out = df2_out[cols_to_use].to_dict('records')

                # Align Rows
                max_len_out = max(len(rows_left_out), len(rows_right_out))
                while len(rows_left_out) < max_len_out: rows_left_out.append({c: "" for c in report_cols})
                while len(rows_right_out) < max_len_out: rows_right_out.append({c: "" for c in report_cols_f2})
                
                # Summary Row
                qty_sum_out = df1_out['Qty'].sum() if not df1_out.empty and 'Qty' in df1_out.columns else 0
                price_sum_out = df1_out['Total Price'].sum() if not df1_out.empty and 'Total Price' in df1_out.columns else 0
                sum_row_left_out = {c: "" for c in report_cols}
                sum_row_left_out['Description'] = f"{f_type}(외주수리)"
                sum_row_left_out['Qty'] = qty_sum_out
                sum_row_left_out['Total Price'] = price_sum_out
                rows_left_out.append(sum_row_left_out)
                
                price_sum_f2_out = df2_out['발주금액'].sum() if not df2_out.empty and '발주금액' in df2_out.columns else 0
                sum_row_right_out = {c: "" for c in report_cols_f2}
                sum_row_right_out['Account name'] = "Total"
                sum_row_right_out['발주금액'] = price_sum_f2_out
                rows_right_out.append(sum_row_right_out)

                # Add 2 Empty Rows
                for _ in range(2):
                    rows_left_out.append({c: "" for c in report_cols})
                    rows_right_out.append({c: "" for c in report_cols_f2})
                
                final_outsourcing_rows_left.extend(rows_left_out)
                final_outsourcing_rows_right.extend(rows_right_out)

            # --- Generate Consolidated Report ---
            file2_dir = os.path.dirname(self.file2_path)
            report_filename = os.path.join(file2_dir, "마감자료 with PRL.xlsx")
            if os.path.exists(report_filename):
                try:
                    with open(report_filename, "a+b"):
                        pass
                except PermissionError:
                    base_name = "마감자료 with PRL"
                    ext = ".xlsx"
                    idx = 1
                    while True:
                        alt_report_filename = os.path.join(file2_dir, f"{base_name}_{idx}{ext}")
                        if not os.path.exists(alt_report_filename):
                            report_filename = alt_report_filename
                            self.log(f"  -> 기존 결과 파일이 열려 있어 새 파일명으로 저장합니다: {os.path.basename(report_filename)}")
                            break
                        idx += 1
            
            df_mat_left = pd.DataFrame(final_material_rows_left, columns=report_cols)
            df_mat_right = pd.DataFrame(final_material_rows_right, columns=report_cols_f2)
            df_out_left = pd.DataFrame(final_outsourcing_rows_left, columns=report_cols)
            df_out_right = pd.DataFrame(final_outsourcing_rows_right, columns=report_cols_f2)
            
            with pd.ExcelWriter(report_filename, engine='openpyxl') as writer:
                df_mat_left.to_excel(writer, sheet_name='자재', index=False, startcol=0)
                df_mat_right.to_excel(writer, sheet_name='자재', index=False, startcol=len(report_cols) + 3)
                df_out_left.to_excel(writer, sheet_name='외주수리', index=False, startcol=0)
                df_out_right.to_excel(writer, sheet_name='외주수리', index=False, startcol=len(report_cols) + 3)
            
            self.log(f"\n[알림] 통합 리포트 '{report_filename}' 생성 완료.")

            # --- Apply Highlighting & Formatting ---
            try:
                wb = openpyxl.load_workbook(report_filename)
                
                # Apply formatting to both sheets
                for sheet_name in ['자재', '외주수리']:
                    if sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        header = [cell.value for cell in ws[1]]

                        # Find columns to format
                        cols_to_format = []
                        if "Unit Price" in header:
                            cols_to_format.append(header.index("Unit Price") + 1)
                        if "Total Price" in header:
                            cols_to_format.append(header.index("Total Price") + 1)
                        if "발주금액" in header:
                            cols_to_format.append(header.index("발주금액") + 1)

                        # Apply number format
                        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                            for col_idx in cols_to_format:
                                cell = row[col_idx - 1]
                                if isinstance(cell.value, (int, float)):
                                    cell.number_format = '#,##0'

                # Apply Highlighting (Material Tab Only)
                if '자재' in wb.sheetnames and missing_doc_nos:
                    ws = wb['자재']
                    header = [cell.value for cell in ws[1]]
                    doc_no_idx = header.index("Doc No.") + 1
                    total_price_idx = header.index("Total Price") + 1

                    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

                    count_highlighted = 0
                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                        doc_no_cell = row[doc_no_idx - 1]
                        val = str(doc_no_cell.value).strip()
                        if val and val in missing_doc_nos:
                            row[total_price_idx - 1].fill = yellow_fill
                            count_highlighted += 1

                    self.log(f"  - '자재' 시트 하이라이트 적용: {count_highlighted} 건")

                # Apply Highlighting (Outsourcing Tab Only)
                if '외주수리' in wb.sheetnames and missing_doc_nos_out:
                    ws = wb['외주수리']
                    header = [cell.value for cell in ws[1]]
                    doc_no_idx = header.index("Doc No.") + 1
                    total_price_idx = header.index("Total Price") + 1
                    
                    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    
                    count_highlighted_out = 0
                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                        doc_no_cell = row[doc_no_idx - 1]
                        val = str(doc_no_cell.value).strip()
                        if val and val in missing_doc_nos_out:
                            row[total_price_idx - 1].fill = yellow_fill
                            count_highlighted_out += 1
                    
                    self.log(f"  - '외주수리' 시트 하이라이트 적용: {count_highlighted_out} 건")

                self.log("  - 숫자 서식(1000단위 콤마) 적용 완료")

                # Apply Highlighting (PR No. Mismatch - Orange)
                # Check both sheets
                orange_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid") # Orange-ish Gold
                
                # Material Sheet
                if '자재' in wb.sheetnames:
                    ws = wb['자재']
                    header = [cell.value for cell in ws[1]]
                    if "PR No." in header:
                        pr_no_idx = header.index("PR No.") + 1
                        count_mismatch = 0
                        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                            pr_no_cell = row[pr_no_idx - 1]
                            pr_no_val = str(pr_no_cell.value).strip() if pr_no_cell.value else ""
                            
                            # Highlight if PR No exists but is NOT in matched set
                            if pr_no_val and pr_no_val not in matched_pr_nos_mat:
                                pr_no_cell.fill = orange_fill
                                count_mismatch += 1
                        self.log(f"  - '자재' 시트 PR No. 미매칭 하이라이트(주황색): {count_mismatch} 건")

                # Outsourcing Sheet
                if '외주수리' in wb.sheetnames:
                    ws = wb['외주수리']
                    header = [cell.value for cell in ws[1]]
                    if "PR No." in header:
                        pr_no_idx = header.index("PR No.") + 1
                        count_mismatch_out = 0
                        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                            pr_no_cell = row[pr_no_idx - 1]
                            pr_no_val = str(pr_no_cell.value).strip() if pr_no_cell.value else ""
                            
                            # Highlight if PR No exists but is NOT in matched set
                            if pr_no_val and pr_no_val not in matched_pr_nos_out:
                                pr_no_cell.fill = orange_fill
                                count_mismatch_out += 1
                        self.log(f"  - '외주수리' 시트 PR No. 미매칭 하이라이트(주황색): {count_mismatch_out} 건")

                # Apply Highlighting (Amount Mismatch - Yellow)
                yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                
                # Material Sheet
                if '자재' in wb.sheetnames:
                    ws = wb['자재']
                    header = [cell.value for cell in ws[1]]
                    if "PR No." in header and "발주금액" in header:
                        pr_no_idx = header.index("PR No.") + 1
                        amount_idx = header.index("발주금액") + 1
                        count_amount_mismatch = 0
                        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                            pr_no_cell = row[pr_no_idx - 1]
                            amount_cell = row[amount_idx - 1]
                            pr_no_val = str(pr_no_cell.value).strip() if pr_no_cell.value else ""
                            
                            # Highlight if PR No is in amount mismatch set
                            if pr_no_val and pr_no_val in amount_mismatch_pr_nos_mat:
                                amount_cell.fill = yellow_fill
                                count_amount_mismatch += 1
                        self.log(f"  - '자재' 시트 금액 불일치 하이라이트(노란색): {count_amount_mismatch} 건")

                # Outsourcing Sheet
                if '외주수리' in wb.sheetnames:
                    ws = wb['외주수리']
                    header = [cell.value for cell in ws[1]]
                    if "PR No." in header and "발주금액" in header:
                        pr_no_idx = header.index("PR No.") + 1
                        amount_idx = header.index("발주금액") + 1
                        count_amount_mismatch_out = 0
                        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                            pr_no_cell = row[pr_no_idx - 1]
                            amount_cell = row[amount_idx - 1]
                            pr_no_val = str(pr_no_cell.value).strip() if pr_no_cell.value else ""
                            
                            # Highlight if PR No is in amount mismatch set
                            if pr_no_val and pr_no_val in amount_mismatch_pr_nos_out:
                                amount_cell.fill = yellow_fill
                                count_amount_mismatch_out += 1
                        self.log(f"  - '외주수리' 시트 금액 불일치 하이라이트(노란색): {count_amount_mismatch_out} 건")

                wb.save(report_filename)
            except Exception as e:
                self.log(f"\n[주의] 엑셀 후처리(서식/하이라이트) 중 오류: {e}")

        except Exception as e:
            self.log(f"\n[치명적 오류] 분석 중 예외 발생: {str(e)}")
            import traceback
            traceback.print_exc()
        finally:
            self.compare_btn.configure(state="normal", text="비교 분석 시작")

if __name__ == "__main__":
    app = ExcelComparatorApp()
    app.mainloop()
